"""Excel exporter for mapping results."""
from typing import List, Optional, Dict, Any
from datetime import datetime
from src.mapping.mapping_lineage import (
    MappingLineage,
    MappingContext,
    MappingType,
)
from src.utils.logger import get_logger

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = get_logger(__name__)


class ExcelExporter:
    """Export mapping results to Excel with detailed comments."""

    def __init__(self, output_path: str):
        """Initialize exporter."""
        self.output_path = output_path
        self.workbook = None
        self.worksheet = None

    def export(self, context: MappingContext) -> bool:
        """Export mapping context to Excel."""
        logger.info(f"Exporting mapping context to {self.output_path}")

        try:
            self.workbook = openpyxl.Workbook()
            self.workbook.remove(self.workbook.active)

            # Create tabs
            self._create_summary_tab(context)
            self._create_one_to_one_tab(context)
            self._create_many_to_one_tab(context)
            self._create_one_to_many_tab(context)
            self._create_execution_order_tab(context)
            self._create_destructive_ops_tab(context)
            self._create_metadata_tab(context)

            self.workbook.save(self.output_path)
            logger.info(f"Successfully exported to {self.output_path}")
            return True

        except ImportError:
            logger.error("openpyxl not installed")
            return False
        except Exception as e:
            logger.error(f"Export failed: {e}")
            return False

    def _create_summary_tab(self, context: MappingContext) -> None:
        """Create summary tab."""
        ws = self.workbook.create_sheet("Summary")
        
        stats = context.get_statistics()
        
        ws['A1'] = "Data Mapping Report"
        ws['A2'] = f"Generated: {datetime.now().isoformat()}"
        
        row = 4
        ws[f'A{row}'] = "Mapping Summary"
        row += 1
        
        ws[f'A{row}'] = "Total Mappings"
        ws[f'B{row}'] = stats['total_mappings']
        row += 1
        
        ws[f'A{row}'] = "1:1 Mappings"
        ws[f'B{row}'] = stats['one_to_one']
        row += 1
        
        ws[f'A{row}'] = "N:1 Mappings (Aggregation)"
        ws[f'B{row}'] = stats['many_to_one']
        row += 1
        
        ws[f'A{row}'] = "1:N Mappings (Decomposition)"
        ws[f'B{row}'] = stats['one_to_many']
        row += 1
        
        ws[f'A{row}'] = "Destructive Mappings"
        ws[f'B{row}'] = stats['destructive_mappings']
        row += 1
        
        ws[f'A{row}'] = "Unmapped Source Fields"
        ws[f'B{row}'] = stats['unmapped_source_fields']
        row += 1
        
        ws[f'A{row}'] = "Unmapped Destination Fields"
        ws[f'B{row}'] = stats['unmapped_destination_fields']
        row += 1
        
        ws[f'A{row}'] = "Circular Dependencies"
        ws[f'B{row}'] = stats['circular_dependencies']

    def _create_one_to_one_tab(self, context: MappingContext) -> None:
        """Create 1:1 mappings tab."""
        ws = self.workbook.create_sheet("1-to-1 Mappings")
        
        headers = [
            "Mapping ID",
            "Source Field",
            "Destination Field",
            "Source Type",
            "Dest Type",
            "Type Conversion",
            "Destructive?",
            "Risk Level",
            "Notes",
        ]
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx).value = header
        
        row = 2
        for lineage in context.get_lineages_by_type(MappingType.ONE_TO_ONE):
            if lineage.source_fields and lineage.destination_field:
                ws.cell(row=row, column=1).value = lineage.mapping_id
                ws.cell(row=row, column=2).value = lineage.source_fields[0].name
                ws.cell(row=row, column=3).value = lineage.destination_field.name
                ws.cell(row=row, column=4).value = lineage.source_fields[0].type_info.base_type.value
                ws.cell(row=row, column=5).value = lineage.destination_field.type_info.base_type.value
                ws.cell(row=row, column=6).value = "Yes" if lineage.transformation_pipeline else "No"
                ws.cell(row=row, column=7).value = "Yes" if lineage.is_destructive else "No"
                ws.cell(row=row, column=8).value = lineage.destruction_type
                ws.cell(row=row, column=9).value = lineage.notes
                
                # Color red if destructive
                if lineage.is_destructive:
                    for col in range(1, 10):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FF0000", end_color="FF0000", fill_type="solid"
                        )
                
                row += 1

    def _create_many_to_one_tab(self, context: MappingContext) -> None:
        """Create N:1 mappings (aggregation) tab."""
        ws = self.workbook.create_sheet("N-to-1 Mappings")
        
        headers = [
            "Mapping ID",
            "Source Fields",
            "Destination Field",
            "Aggregation Strategy",
            "Expression",
            "Destructive?",
            "Data Loss Description",
            "Notes",
        ]
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx).value = header
        
        row = 2
        for lineage in context.get_lineages_by_type(MappingType.MANY_TO_ONE):
            if lineage.aggregation_rule:
                source_names = ", ".join(f.name for f in lineage.source_fields)
                
                ws.cell(row=row, column=1).value = lineage.mapping_id
                ws.cell(row=row, column=2).value = source_names
                ws.cell(row=row, column=3).value = lineage.destination_field.name if lineage.destination_field else "N/A"
                ws.cell(row=row, column=4).value = lineage.aggregation_rule.strategy.value
                ws.cell(row=row, column=5).value = lineage.aggregation_rule.aggregation_expression
                ws.cell(row=row, column=6).value = "Yes" if lineage.is_destructive else "No"
                ws.cell(row=row, column=7).value = lineage.aggregation_rule.data_loss_description
                ws.cell(row=row, column=8).value = lineage.notes
                
                # Color yellow if destructive
                if lineage.is_destructive:
                    for col in range(1, 9):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                        )
                
                row += 1

    def _create_one_to_many_tab(self, context: MappingContext) -> None:
        """Create 1:N mappings (decomposition) tab."""
        ws = self.workbook.create_sheet("1-to-N Mappings")
        
        headers = [
            "Mapping ID",
            "Source Field",
            "Destination Fields",
            "Decomposition Strategy",
            "Expression",
            "Destructive?",
            "Data Loss Description",
            "Notes",
        ]
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx).value = header
        
        row = 2
        for lineage in context.get_lineages_by_type(MappingType.ONE_TO_MANY):
            if lineage.decomposition_rule:
                dest_names = ", ".join(f.name for f in lineage.destination_fields)
                
                ws.cell(row=row, column=1).value = lineage.mapping_id
                ws.cell(row=row, column=2).value = lineage.source_fields[0].name if lineage.source_fields else "N/A"
                ws.cell(row=row, column=3).value = dest_names
                ws.cell(row=row, column=4).value = lineage.decomposition_rule.strategy.value
                ws.cell(row=row, column=5).value = lineage.decomposition_rule.decomposition_expression
                ws.cell(row=row, column=6).value = "Yes" if lineage.is_destructive else "No"
                ws.cell(row=row, column=7).value = lineage.decomposition_rule.data_loss_description
                ws.cell(row=row, column=8).value = lineage.notes
                
                # Color yellow if destructive
                if lineage.is_destructive:
                    for col in range(1, 9):
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                        )
                
                row += 1

    def _create_execution_order_tab(self, context: MappingContext) -> None:
        """Create execution order tab."""
        ws = self.workbook.create_sheet("Execution Order")
        
        headers = ["Sequence", "Mapping ID", "Mapping Type", "Sources", "Destination", "Dependencies"]
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx).value = header
        
        row = 2
        for idx, mapping_id in enumerate(context.execution_order, 1):
            lineage = context.get_lineage_by_id(mapping_id)
            if lineage:
                sources = ", ".join(f.name for f in lineage.source_fields) if lineage.source_fields else "N/A"
                dest = lineage.destination_field.name if lineage.destination_field else lineage.destination_fields[0].name if lineage.destination_fields else "N/A"
                deps = ", ".join(lineage.depends_on) if lineage.depends_on else "None"
                
                ws.cell(row=row, column=1).value = idx
                ws.cell(row=row, column=2).value = mapping_id
                ws.cell(row=row, column=3).value = lineage.mapping_type.value
                ws.cell(row=row, column=4).value = sources
                ws.cell(row=row, column=5).value = dest
                ws.cell(row=row, column=6).value = deps
                
                row += 1

    def _create_destructive_ops_tab(self, context: MappingContext) -> None:
        """Create destructive operations tab."""
        ws = self.workbook.create_sheet("Destructive Operations")
        
        headers = [
            "Mapping ID",
            "Type",
            "Severity",
            "Sources",
            "Destination",
            "Details",
            "Recommendation",
        ]
        
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx).value = header
        
        row = 2
        for lineage in context.get_destructive_lineages():
            sources = ", ".join(f.name for f in lineage.source_fields) if lineage.source_fields else "N/A"
            dest = lineage.destination_field.name if lineage.destination_field else lineage.destination_fields[0].name if lineage.destination_fields else "N/A"
            
            ws.cell(row=row, column=1).value = lineage.mapping_id
            ws.cell(row=row, column=2).value = lineage.mapping_type.value
            ws.cell(row=row, column=3).value = lineage.destruction_type
            ws.cell(row=row, column=4).value = sources
            ws.cell(row=row, column=5).value = dest
            ws.cell(row=row, column=6).value = lineage.destruction_details
            ws.cell(row=row, column=7).value = lineage.notes
            
            row += 1

    def _create_metadata_tab(self, context: MappingContext) -> None:
        """Create metadata tab."""
        ws = self.workbook.create_sheet("Metadata")
        
        ws['A1'] = "Context ID"
        ws['B1'] = context.context_id
        
        ws['A2'] = "Generated"
        ws['B2'] = datetime.now().isoformat()
        
        ws['A4'] = "Source Schemas"
        row = 5
        for name in context.source_schemas.keys():
            ws[f'A{row}'] = name
            row += 1
        
        ws['C4'] = "Destination Schemas"
        row = 5
        for name in context.destination_schemas.keys():
            ws[f'C{row}'] = name
            row += 1
